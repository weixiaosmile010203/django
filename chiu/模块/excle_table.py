from openpyxl import load_workbook
import os


#       取得工作表
wb = load_workbook(r'C:\Users\Chiu\Desktop\test.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet1']
print(sheet['A4'].value)#A4单元格的值

print(sheet.max_row)#最大行数

print(sheet.max_column)#最大列数

for i in sheet['A']: #显示A列的所有单元格
    print(i.value, end='')

for i in sheet['1']:
    print(i.value, end='') #显示1行所有单元格